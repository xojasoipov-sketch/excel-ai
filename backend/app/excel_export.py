"""Turns an AI-generated formula string into a real, openable .xlsx workbook:
a small sample table filled with plausible data plus the *live* formula written
into a cell (Excel recalculates it the moment the file is opened). A bare
formula string is only useful if you already have exactly matching column
letters in your own sheet — handing back a real file people can immediately
open, see working, and adapt is far more useful.

Works on *any* formula, not just the 24 curated Formula Library templates —
it parses whatever cell/range references and quoted criteria the model used,
so this covers ad-hoc AI answers from the fileless chat (web and bot alike).
"""
import re
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter

_CELL_RANGE_RE = re.compile(r"\$?([A-Za-z]{1,2})\$?(\d{1,7})(?::\$?([A-Za-z]{1,2})\$?(\d{1,7}))?")
# Whole-column ranges (A:A, B:D) — the form models reach for most often. The
# lookarounds keep it from matching the middle of a row-bounded range like A1:B5.
_WHOLE_COLUMN_RE = re.compile(r"(?<![A-Za-z0-9$])\$?([A-Za-z]{1,2})\$?:\$?([A-Za-z]{1,2})\$?(?![A-Za-z0-9])")
_QUOTED_RE = re.compile(r'"([^"]*)"')
# A numeric comparison criteria such as ">100" or "<=5.5"
_COMPARISON_RE = re.compile(r"^(>=|<=|<>|>|<|=)\s*(-?\d+(?:\.\d+)?)$")
# A range immediately followed by a comma then a quoted literal, e.g.
# B1:B10,"Toshkent" or B:B,"Toshkent" — that column is the one being matched on.
_CRITERIA_PAIR_RE = re.compile(
    r"\$?([A-Za-z]{1,2})\$?(?:\d{1,7})?:\$?[A-Za-z]{1,2}\$?(?:\d{1,7})?\s*,\s*\"([^\"]*)\""
)

DEFAULT_SAMPLE_ROWS = 10  # used when the formula names no explicit row numbers

_FALLBACK_LABELS = ["Toshkent", "Andijon", "Buxoro", "Namangan", "Farg'ona"]
_HEADER_FONT = Font(bold=True)
_RESULT_FONT = Font(bold=True, color="087B40")


def _extract_formula(text: str) -> Optional[str]:
    """Pulls the formula out of a ```excel ... ``` (or any ``` ... ```) code
    block, which the system prompts always ask the model to use. Falls back to
    the first line that starts with '=' if no fenced block is present."""
    fenced = re.search(r"```(?:excel|formula)?\s*\n?(.*?)```", text, re.DOTALL | re.IGNORECASE)
    candidate = fenced.group(1).strip() if fenced else None
    if not candidate:
        for line in text.splitlines():
            stripped = line.strip().strip("`")
            if stripped.startswith("="):
                candidate = stripped
                break
    if not candidate:
        return None
    # A code block occasionally contains more than one line; the formula is
    # whichever line actually starts with '='.
    for line in candidate.splitlines():
        stripped = line.strip()
        if stripped.startswith("="):
            return stripped
    return candidate if candidate.startswith("=") else None


def build_workbook_for_formula(ai_response_text: str) -> Optional[bytes]:
    """Returns .xlsx bytes with sample data + the live formula, or None if no
    formula could be found in the AI's response (e.g. it just asked a
    clarifying question instead)."""
    formula = _extract_formula(ai_response_text)
    if not formula:
        return None

    columns = set()
    max_row = 1

    def add_span(letter_a: str, letter_b: str) -> None:
        """Fill every column the range spans (B:D means B, C and D), not just the
        two letters that literally appear in the formula text."""
        idx_a = column_index_from_string(letter_a.upper())
        idx_b = column_index_from_string(letter_b.upper())
        for idx in range(min(idx_a, idx_b), max(idx_a, idx_b) + 1):
            columns.add(get_column_letter(idx))

    for m in _CELL_RANGE_RE.finditer(formula):
        col1, row1, col2, row2 = m.groups()
        max_row = max(max_row, int(row1))
        if col2:
            add_span(col1, col2)
            max_row = max(max_row, int(row2))
        else:
            columns.add(get_column_letter(column_index_from_string(col1.upper())))

    # Whole-column ranges (=SUMIF(B:B,"Toshkent",A:A)) carry no row numbers, so
    # they'd otherwise yield no columns at all and skip the workbook entirely.
    for m in _WHOLE_COLUMN_RE.finditer(formula):
        add_span(*m.groups())

    if not columns:
        return None
    if max_row == 1:
        max_row = DEFAULT_SAMPLE_ROWS
    max_row = min(max_row, 40)  # keep the sample file small

    # Text criteria (B:B,"Toshkent") become labels in that column. Numeric
    # comparisons (C:C,">100") must NOT be written in literally — the cells need
    # numbers straddling the threshold, otherwise the condition matches nothing
    # and the sample file doesn't actually demonstrate the formula.
    criteria_columns: dict[str, str] = {}
    threshold_columns: dict[str, float] = {}
    for m in _CRITERIA_PAIR_RE.finditer(formula):
        col, literal = m.groups()
        comparison = _COMPARISON_RE.match(literal.strip())
        if comparison:
            threshold_columns[col.upper()] = float(comparison.group(2))
        else:
            criteria_columns[col.upper()] = literal

    other_literals = [
        v for v in _QUOTED_RE.findall(formula)
        if v not in criteria_columns.values() and not _COMPARISON_RE.match(v.strip())
    ]
    label_pool = (other_literals + _FALLBACK_LABELS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Namuna"

    sorted_columns = sorted(columns, key=column_index_from_string)

    for row in range(1, max_row + 1):
        for col in sorted_columns:
            cell = ws[f"{col}{row}"]
            if col in criteria_columns:
                # Alternate between the model's own literal and a couple of
                # plausible alternates, so the formula's condition genuinely
                # matches on some rows and not others when opened in Excel.
                cell.value = criteria_columns[col] if row % 2 == 1 else label_pool[row % len(label_pool)]
            elif col in threshold_columns:
                # Straddle the threshold: odd rows clear it, even rows don't.
                threshold = threshold_columns[col]
                cell.value = round(threshold + 25 + row if row % 2 == 1 else max(threshold - 25 - row, 0), 2)
            else:
                cell.value = (row * 37 % 460) + 15

    result_row = max_row + 2
    result_label_col = sorted_columns[0]
    label_idx = column_index_from_string(result_label_col)
    formula_idx = label_idx + 1 if len(sorted_columns) == 1 else column_index_from_string(sorted_columns[-1]) + 1

    label_cell = ws.cell(row=result_row, column=label_idx, value="Natija:")
    label_cell.font = _HEADER_FONT
    result_cell = ws.cell(row=result_row, column=formula_idx, value=formula)
    result_cell.font = _RESULT_FONT

    for col in sorted_columns:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions[get_column_letter(formula_idx)].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
